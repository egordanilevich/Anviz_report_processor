# -*- coding: utf-8 -*-
"""
Created on Fri Feb 20 03:26:09 2026

@author: Egor
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import os
from datetime import datetime, time
import xlrd
from xlrd import xldate_as_tuple
import openpyxl
from openpyxl.styles import PatternFill

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Обработчик данных СКУД")
        self.root.geometry("1280x720")
        self.root.resizable(False, False)

        self.input_file_path = tk.StringVar()
        self.output_file_path = tk.StringVar()

        # Параметры рабочего дня
        self.work_start = tk.StringVar(value="09:00")
        self.work_end = tk.StringVar(value="18:00")
        self.work_hours_norm = tk.StringVar(value="8.0")

        self.create_widgets()

    def create_widgets(self):
        # Входной файл
        tk.Label(self.root, text="Входной файл:").grid(row=0, column=0, padx=5, pady=10, sticky="w")
        tk.Entry(self.root, textvariable=self.input_file_path, width=60).grid(row=0, column=1, padx=5, pady=10)
        tk.Button(self.root, text="Обзор...", command=self.select_input_file).grid(row=0, column=2, padx=5, pady=10)

        # Выходной файл
        tk.Label(self.root, text="Выходной файл:").grid(row=1, column=0, padx=5, pady=10, sticky="w")
        tk.Entry(self.root, textvariable=self.output_file_path, width=60).grid(row=1, column=1, padx=5, pady=10)
        tk.Button(self.root, text="Сохранить как...", command=self.select_output_file).grid(row=1, column=2, padx=5, pady=10)

        # Параметры
        tk.Label(self.root, text="Время начала рабочего дня (ЧЧ:ММ):").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(self.root, textvariable=self.work_start, width=10).grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(self.root, text="Время окончания рабочего дня (ЧЧ:ММ):").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(self.root, textvariable=self.work_end, width=10).grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(self.root, text="Норма часов работы (часов):").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(self.root, textvariable=self.work_hours_norm, width=10).grid(row=4, column=1, padx=5, pady=5, sticky="w")

        # Кнопка обработки
        tk.Button(self.root, text="Обработать", command=self.process_files, width=20, height=2).grid(row=5, column=1, pady=20)

    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Выберите входной файл",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.input_file_path.set(filename)

    def select_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Сохранить выходной файл как",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file_path.set(filename)

    def process_files(self):
        input_path = self.input_file_path.get()
        output_path = self.output_file_path.get()

        if not input_path or not output_path:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите входной и выходной файлы.")
            return

        try:
            work_start = datetime.strptime(self.work_start.get(), "%H:%M").time()
            work_end = datetime.strptime(self.work_end.get(), "%H:%M").time()
            norm_hours = float(self.work_hours_norm.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат времени или нормы часов.\nИспользуйте ЧЧ:ММ и число (например, 8.0).")
            return

        try:
            self.parse_xls_to_result(input_path, output_path, work_start, work_end, norm_hours)
            messagebox.showinfo("Успех", f"Обработка завершена!\nРезультат сохранён в:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка:\n{str(e)}")

    def parse_xls_to_result(self, input_path, output_path, work_start, work_end, norm_hours):
        # Открываем исходный .xls файл через xlrd
        wb = xlrd.open_workbook(input_path)
        sheet = wb.sheet_by_index(0)  # первый лист

        # Создаём новую книгу для результатов
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Результат"

        columns = [
            "фио", "табельный номер", "Отдел", "Должность",
            "дата", "время входа на работу", "время выхода с работы",
            "входы выходы количество", "Опаздания", "Рабочее время",
            "Время отсутствия", "Переработка"
        ]
        ws_out.append(columns)

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        row_idx = 0
        total_rows = sheet.nrows

        while row_idx < total_rows:
            cell_a = sheet.cell_value(row_idx, 0)
            cell_b = sheet.cell_value(row_idx, 1)
            cell_e = sheet.cell_value(row_idx, 4)

            # Поиск строки с данными сотрудника
            if isinstance(cell_a, (int, float)) and cell_b and isinstance(cell_b, str) and cell_b.strip():
                tab_number = str(int(cell_a)) if isinstance(cell_a, (int, float)) else str(cell_a)
                fio = cell_b.strip()
                department = cell_e.strip() if cell_e else ""

                data_row = row_idx + 1
                while data_row < total_rows:
                    # Проверка на начало следующего сотрудника
                    next_a = sheet.cell_value(data_row, 0)
                    next_b = sheet.cell_value(data_row, 1)
                    if isinstance(next_a, (int, float)) and next_b and isinstance(next_b, str) and next_b.strip():
                        break

                    # Пропуск строки "Составитель"
                    cell_h = sheet.cell_value(data_row, 7)  # колонка H
                    if cell_h and "Составитель" in str(cell_h):
                        data_row += 1
                        continue

                    # Обработка даты
                    date_val = sheet.cell_value(data_row, 0)
                    date_obj = None
                    if date_val:
                        if sheet.cell_type(data_row, 0) == xlrd.XL_CELL_DATE:
                            date_tuple = xldate_as_tuple(date_val, wb.datemode)
                            date_obj = datetime(*date_tuple[:3])  # только дата
                        elif isinstance(date_val, str):
                            try:
                                date_obj = datetime.strptime(date_val.strip(), "%Y-%m-%d")
                            except:
                                pass

                    if date_obj:
                        time_str = sheet.cell_value(data_row, 2)  # колонка C
                        count_val = sheet.cell_value(data_row, 7)  # колонка H
                        worked_val = sheet.cell_value(data_row, 9)  # колонка J

                        try:
                            count = int(float(count_val)) if count_val else 0
                        except:
                            count = 0
                        try:
                            worked = float(worked_val) if worked_val else 0.0
                        except:
                            worked = 0.0

                        entry_time_str = ""
                        exit_time_str = ""
                        entry_time_obj = None
                        exit_time_obj = None

                        if time_str and isinstance(time_str, str):
                            times = [t.strip() for t in time_str.split('-') if t.strip()]
                            if times:
                                time_objects = []
                                for t in times:
                                    try:
                                        t_obj = datetime.strptime(t, "%H:%M:%S").time()
                                        time_objects.append(t_obj)
                                    except:
                                        pass
                                if time_objects:
                                    first = min(time_objects)
                                    last = max(time_objects)
                                    entry_time_obj = first
                                    exit_time_obj = last
                                    entry_time_str = first.strftime("%H:%M")
                                    exit_time_str = last.strftime("%H:%M")

                        # Опоздание
                        delay = ""
                        if entry_time_obj and entry_time_obj > work_start:
                            delta = datetime.combine(date_obj, entry_time_obj) - datetime.combine(date_obj, work_start)
                            minutes = delta.seconds // 60
                            delay = f"{minutes} мин"

                        # Ранний уход
                        early_leave = ""
                        if exit_time_obj and exit_time_obj < work_end:
                            delta = datetime.combine(date_obj, work_end) - datetime.combine(date_obj, exit_time_obj)
                            minutes = delta.seconds // 60
                            early_leave = f"{minutes} мин"

                        # Переработка/недоработка
                        excess = ""
                        if worked:
                            diff = worked - norm_hours
                            if diff > 0.01:
                                excess = f"+{diff:.2f}"
                            elif diff < -0.01:
                                excess = f"{diff:.2f}"

                        row_data = [
                            fio,
                            tab_number,
                            department,
                            "",
                            date_obj.strftime("%d.%m.%Y"),
                            entry_time_str,
                            exit_time_str,
                            count,
                            delay,
                            worked,
                            early_leave,
                            excess
                        ]
                        ws_out.append(row_data)

                        # Заливка
                        current_row = ws_out.max_row
                        if delay:
                            ws_out.cell(row=current_row, column=6).fill = red_fill
                        if early_leave:
                            ws_out.cell(row=current_row, column=7).fill = red_fill

                    data_row += 1

                row_idx = data_row
            else:
                row_idx += 1

        wb_out.save(output_path)
        wb_out.close()

        if ws_out.max_row == 1:
            raise Exception("Не найдено данных сотрудников в файле.")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()