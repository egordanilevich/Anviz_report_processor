# -*- coding: utf-8 -*-
"""
Created on Fri Feb 20 03:33:49 2026

@author: Egor
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime, time
import threading
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Обработчик данных СКУД")
        self.root.minsize(600, 300)

        self.input_file_path = tk.StringVar()
        self.output_file_path = tk.StringVar()

        self.work_start = tk.StringVar(value="09:00")
        self.work_end = tk.StringVar(value="18:00")
        self.work_hours_norm = tk.StringVar(value="8.0")

        self.progress = None
        self.process_button = None
        self.create_widgets()

        # Настройка весов для адаптивного расширения
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(5, weight=1)  # чтобы прогресс-бар не сжимался

    def create_widgets(self):
        # Входной файл
        tk.Label(self.root, text="Входной файл:").grid(row=0, column=0, padx=5, pady=10, sticky="w")
        tk.Entry(self.root, textvariable=self.input_file_path).grid(row=0, column=1, padx=5, pady=10, sticky="ew")
        tk.Button(self.root, text="Обзор...", command=self.select_input_file).grid(row=0, column=2, padx=5, pady=10)

        # Выходной файл
        tk.Label(self.root, text="Выходной файл:").grid(row=1, column=0, padx=5, pady=10, sticky="w")
        tk.Entry(self.root, textvariable=self.output_file_path).grid(row=1, column=1, padx=5, pady=10, sticky="ew")
        tk.Button(self.root, text="Сохранить как...", command=self.select_output_file).grid(row=1, column=2, padx=5, pady=10)

        # Параметры
        tk.Label(self.root, text="Время начала рабочего дня (ЧЧ:ММ):").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(self.root, textvariable=self.work_start, width=10).grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(self.root, text="Время окончания рабочего дня (ЧЧ:ММ):").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(self.root, textvariable=self.work_end, width=10).grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(self.root, text="Норма часов работы (часов):").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(self.root, textvariable=self.work_hours_norm, width=10).grid(row=4, column=1, padx=5, pady=5, sticky="w")

        # Кнопка обработки
        self.process_button = tk.Button(self.root, text="Обработать", command=self.start_processing, width=20, height=2)
        self.process_button.grid(row=5, column=1, pady=20)

        # Прогресс-бар (изначально скрыт)
        self.progress = ttk.Progressbar(self.root, mode='indeterminate')
        self.progress.grid(row=6, column=0, columnspan=3, padx=20, pady=10, sticky="ew")
        self.progress.grid_remove()  # скрываем

    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Выберите входной файл",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.input_file_path.set(filename)

    def select_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Сохранить выходной файл как",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file_path.set(filename)

    def start_processing(self):
        """Запускает обработку в отдельном потоке, чтобы не блокировать GUI."""
        input_path = self.input_file_path.get()
        output_path = self.output_file_path.get()

        if not input_path or not output_path:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите входной и выходной файлы.")
            return

        try:
            work_start = datetime.strptime(self.work_start.get(), "%H:%M").time()
            work_end = datetime.strptime(self.work_end.get(), "%H:%M").time()
            norm_hours = float(self.work_hours_norm.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат времени или нормы часов.\nИспользуйте ЧЧ:ММ и число (например, 8.0).")
            return

        # Отключаем кнопку и показываем прогресс
        self.process_button.config(state=tk.DISABLED)
        self.progress.grid()
        self.progress.start(10)  # интервал 10 мс

        # Запускаем поток
        thread = threading.Thread(target=self.process_files, args=(input_path, output_path, work_start, work_end, norm_hours))
        thread.daemon = True
        thread.start()

    def process_files(self, input_path, output_path, work_start, work_end, norm_hours):
        """Выполняется в потоке."""
        try:
            self.parse_xls_to_result(input_path, output_path, work_start, work_end, norm_hours)
            # Сообщение об успехе в главном потоке
            self.root.after(0, self.on_success, output_path)
        except Exception as e:
            self.root.after(0, self.on_error, str(e))

    def on_success(self, output_path):
        """Вызывается после успешной обработки."""
        self.progress.stop()
        self.progress.grid_remove()
        self.process_button.config(state=tk.NORMAL)
        messagebox.showinfo("Успех", f"Обработка завершена!\nРезультат сохранён в:\n{output_path}")

    def on_error(self, error_msg):
        """Вызывается при ошибке."""
        self.progress.stop()
        self.progress.grid_remove()
        self.process_button.config(state=tk.NORMAL)
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n{error_msg}")

    # Вся логика parse_xls_to_result остаётся без изменений
    def parse_xls_to_result(self, input_path, output_path, work_start, work_end, norm_hours):
        # Полный код функции из предыдущего ответа (без изменений)
        # Вставляем сюда ту же реализацию
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb.active

        results = []
        columns = [
            "фио", "табельный номер", "Отдел", "Должность",
            "дата", "время входа на работу", "время выхода с работы",
            "входы выходы количество", "Опаздания", "Рабочее время",
            "Время отсутствия", "Переработка"
        ]

        row_idx = 1
        while row_idx <= ws.max_row:
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value
            cell_e = ws.cell(row=row_idx, column=5).value

            if isinstance(cell_a, (int, float)) and cell_b and isinstance(cell_b, str) and cell_b.strip():
                tab_number = str(int(cell_a)) if isinstance(cell_a, (int, float)) else str(cell_a)
                fio = cell_b.strip()
                department = cell_e.strip() if cell_e else ""

                data_row = row_idx + 1
                while data_row <= ws.max_row:
                    next_a = ws.cell(row=data_row, column=1).value
                    next_b = ws.cell(row=data_row, column=2).value
                    if isinstance(next_a, (int, float)) and next_b and isinstance(next_b, str) and next_b.strip():
                        break

                    if ws.cell(row=data_row, column=8).value and "Составитель" in str(ws.cell(row=data_row, column=8).value):
                        data_row += 1
                        continue

                    date_val = ws.cell(row=data_row, column=1).value
                    date_obj = None
                    if isinstance(date_val, datetime):
                        date_obj = date_val
                    elif isinstance(date_val, str):
                        try:
                            date_obj = datetime.strptime(date_val.strip(), "%Y-%m-%d")
                        except:
                            pass

                    if date_obj:
                        time_str = ws.cell(row=data_row, column=3).value
                        count_val = ws.cell(row=data_row, column=8).value
                        worked_val = ws.cell(row=data_row, column=10).value

                        try:
                            count = int(float(count_val)) if count_val is not None else 0
                        except:
                            count = 0
                        try:
                            worked = float(worked_val) if worked_val is not None else 0.0
                        except:
                            worked = 0.0

                        entry_time_str = ""
                        exit_time_str = ""
                        entry_time_obj = None
                        exit_time_obj = None
                        if time_str and isinstance(time_str, str):
                            times = [t.strip() for t in time_str.split('-') if t.strip()]
                            if times:
                                time_objects = []
                                for t in times:
                                    try:
                                        t_obj = datetime.strptime(t, "%H:%M:%S").time()
                                        time_objects.append(t_obj)
                                    except:
                                        pass
                                if time_objects:
                                    first = min(time_objects)
                                    last = max(time_objects)
                                    entry_time_obj = first
                                    exit_time_obj = last
                                    entry_time_str = first.strftime("%H:%M")
                                    exit_time_str = last.strftime("%H:%M")

                        delay = ""
                        if entry_time_obj and entry_time_obj > work_start:
                            delta = datetime.combine(date_obj, entry_time_obj) - datetime.combine(date_obj, work_start)
                            minutes = delta.seconds // 60
                            delay = f"{minutes} мин"

                        early_leave = ""
                        if exit_time_obj and exit_time_obj < work_end:
                            delta = datetime.combine(date_obj, work_end) - datetime.combine(date_obj, exit_time_obj)
                            minutes = delta.seconds // 60
                            early_leave = f"{minutes} мин"

                        excess = ""
                        if worked:
                            diff = worked - norm_hours
                            if diff > 0.01:
                                excess = f"+{diff:.2f}"
                            elif diff < -0.01:
                                excess = f"{diff:.2f}"

                        record = {
                            "фио": fio,
                            "табельный номер": tab_number,
                            "Отдел": department,
                            "Должность": "",
                            "дата": date_obj.strftime("%d.%m.%Y"),
                            "время входа на работу": entry_time_str,
                            "время выхода с работы": exit_time_str,
                            "входы выходы количество": count,
                            "Опаздания": delay,
                            "Рабочее время": worked,
                            "Время отсутствия": early_leave,
                            "Переработка": excess
                        }
                        results.append(record)

                    data_row += 1

                row_idx = data_row
            else:
                row_idx += 1

        wb.close()

        if not results:
            raise Exception("Не найдено данных сотрудников в файле.")

        df_result = pd.DataFrame(results, columns=columns)

        temp_output = output_path + ".temp.xlsx"
        df_result.to_excel(temp_output, index=False)

        wb_out = openpyxl.load_workbook(temp_output)
        ws_out = wb_out.active

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        col_idx_entry = None
        col_idx_exit = None
        for col in range(1, ws_out.max_column + 1):
            cell = ws_out.cell(row=1, column=col)
            if cell.value == "время входа на работу":
                col_idx_entry = col
            elif cell.value == "время выхода с работы":
                col_idx_exit = col

        for row in range(2, ws_out.max_row + 1):
            delay_cell = ws_out.cell(row, col_idx_entry)
            delay_val = ws_out.cell(row, columns.index("Опаздания") + 1).value
            if delay_val and "мин" in str(delay_val):
                delay_cell.fill = red_fill

            early_cell = ws_out.cell(row, col_idx_exit)
            early_val = ws_out.cell(row, columns.index("Время отсутствия") + 1).value
            if early_val and "мин" in str(early_val):
                early_cell.fill = red_fill

        wb_out.save(output_path)
        wb_out.close()
        os.remove(temp_output)

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()